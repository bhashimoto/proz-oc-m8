from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHelper:
    """Utility wrapper around openpyxl for reading, writing, and manipulating Excel files."""

    def __init__(self, file_path: str | Path | None = None):
        self.file_path = Path(file_path) if file_path else None
        if file_path and self.file_path.exists():
            self.workbook = openpyxl.load_workbook(self.file_path)
        else:
            self.workbook = Workbook()
            # Remove the default empty sheet when creating without a path
            if not file_path:
                self.workbook.active.title = "Sheet1"

    # ------------------------------------------------------------------
    # Workbook / sheet management
    # ------------------------------------------------------------------

    def create_sheet(self, title: str, position: int | None = None) -> Worksheet:
        """Add a new sheet to the workbook."""
        return self.workbook.create_sheet(title=title, index=position)

    def get_sheet(self, name: str) -> Worksheet:
        """Return a sheet by name."""
        return self.workbook[name]

    def get_active_sheet(self) -> Worksheet:
        """Return the currently active sheet."""
        return self.workbook.active

    def set_active_sheet(self, name: str) -> None:
        """Set the active sheet by name."""
        self.workbook.active = self.workbook[name]

    def delete_sheet(self, name: str) -> None:
        """Remove a sheet from the workbook."""
        del self.workbook[name]

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """Rename an existing sheet."""
        self.workbook[old_name].title = new_name

    def sheet_names(self) -> list[str]:
        """Return all sheet names in the workbook."""
        return self.workbook.sheetnames

    # ------------------------------------------------------------------
    # Reading
    # ------------------------------------------------------------------

    def read_cell(self, row: int, col: int | str, sheet: str | None = None) -> Any:
        """Read a single cell value. col can be an int (1-based) or a letter ('A')."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        return ws.cell(row=row, column=col_idx).value

    def read_row(self, row: int, sheet: str | None = None) -> list[Any]:
        """Return all cell values in a row as a list."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        return [cell.value for cell in ws[row]]

    def read_column(self, col: int | str, sheet: str | None = None) -> list[Any]:
        """Return all cell values in a column as a list."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        return [cell.value for row in ws.iter_rows(min_col=col_idx, max_col=col_idx) for cell in row]

    def read_range(
        self,
        min_row: int,
        min_col: int | str,
        max_row: int,
        max_col: int | str,
        sheet: str | None = None,
    ) -> list[list[Any]]:
        """Return a 2-D list of values from the given cell range."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        min_c = column_index_from_string(min_col) if isinstance(min_col, str) else min_col
        max_c = column_index_from_string(max_col) if isinstance(max_col, str) else max_col
        return [
            [cell.value for cell in row]
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_c, max_col=max_c)
        ]

    def read_all(self, sheet: str | None = None, header: bool = True) -> list[dict | list]:
        """
        Read all data from a sheet.

        When header=True, returns a list of dicts keyed by the first row values.
        When header=False, returns a list of lists.
        """
        ws = self.workbook[sheet] if sheet else self.workbook.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        if not header:
            return [list(row) for row in rows]
        keys = rows[0]
        return [dict(zip(keys, row)) for row in rows[1:]]

    # ------------------------------------------------------------------
    # Writing
    # ------------------------------------------------------------------

    def write_cell(
        self,
        row: int,
        col: int | str,
        value: Any,
        sheet: str | None = None,
    ) -> None:
        """Write a value to a single cell."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.cell(row=row, column=col_idx, value=value)

    def write_row(
        self,
        row: int,
        data: list[Any],
        start_col: int | str = 1,
        sheet: str | None = None,
    ) -> None:
        """Write a list of values into consecutive cells in a row."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        start = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
        for offset, value in enumerate(data):
            ws.cell(row=row, column=start + offset, value=value)

    def write_column(
        self,
        col: int | str,
        data: list[Any],
        start_row: int = 1,
        sheet: str | None = None,
    ) -> None:
        """Write a list of values into consecutive cells in a column."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        for offset, value in enumerate(data):
            ws.cell(row=start_row + offset, column=col_idx, value=value)

    def write_data(
        self,
        data: list[list[Any]] | list[dict],
        start_row: int = 1,
        start_col: int | str = 1,
        write_header: bool = True,
        sheet: str | None = None,
    ) -> None:
        """
        Write a collection of rows to the sheet.

        Accepts a list of lists or a list of dicts. When dicts are provided
        and write_header=True, dict keys are written as the header row.
        """
        ws = self.workbook[sheet] if sheet else self.workbook.active
        start_c = column_index_from_string(start_col) if isinstance(start_col, str) else start_col
        current_row = start_row

        if data and isinstance(data[0], dict):
            if write_header:
                headers = list(data[0].keys())
                for offset, h in enumerate(headers):
                    ws.cell(row=current_row, column=start_c + offset, value=h)
                current_row += 1
            for record in data:
                for offset, value in enumerate(record.values()):
                    ws.cell(row=current_row, column=start_c + offset, value=value)
                current_row += 1
        else:
            for row_data in data:
                for offset, value in enumerate(row_data):
                    ws.cell(row=current_row, column=start_c + offset, value=value)
                current_row += 1

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    def apply_font(
        self,
        row: int,
        col: int | str,
        bold: bool = False,
        italic: bool = False,
        size: int | None = None,
        color: str | None = None,
        sheet: str | None = None,
    ) -> None:
        """Apply font styling to a cell."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(
            bold=bold,
            italic=italic,
            size=size,
            color=color,
        )

    def apply_fill(
        self,
        row: int,
        col: int | str,
        color: str,
        sheet: str | None = None,
    ) -> None:
        """Apply a solid background fill to a cell. color is an ARGB hex string, e.g. 'FFFF0000'."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.cell(row=row, column=col_idx).fill = PatternFill(
            fill_type="solid", fgColor=color
        )

    def apply_alignment(
        self,
        row: int,
        col: int | str,
        horizontal: str = "general",
        vertical: str = "bottom",
        wrap_text: bool = False,
        sheet: str | None = None,
    ) -> None:
        """Apply alignment to a cell."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.cell(row=row, column=col_idx).alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrap_text=wrap_text
        )

    def apply_border(
        self,
        row: int,
        col: int | str,
        style: str = "thin",
        sheet: str | None = None,
    ) -> None:
        """Apply a uniform border around a cell."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        side = Side(style=style)
        ws.cell(row=row, column=col_idx).border = Border(
            left=side, right=side, top=side, bottom=side
        )

    def set_column_width(
        self, col: int | str, width: float, sheet: str | None = None
    ) -> None:
        """Set the width of a column."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_letter = get_column_letter(col) if isinstance(col, int) else col
        ws.column_dimensions[col_letter].width = width

    def set_row_height(self, row: int, height: float, sheet: str | None = None) -> None:
        """Set the height of a row."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        ws.row_dimensions[row].height = height

    def auto_fit_columns(self, sheet: str | None = None) -> None:
        """Adjust each column width to fit its longest cell value."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        for col_cells in ws.columns:
            max_length = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=0,
            )
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

    # ------------------------------------------------------------------
    # Formulas & merging
    # ------------------------------------------------------------------

    def write_formula(
        self,
        row: int,
        col: int | str,
        formula: str,
        sheet: str | None = None,
    ) -> None:
        """Write an Excel formula string (e.g. '=SUM(A1:A10)') to a cell."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.cell(row=row, column=col_idx, value=formula)

    def merge_cells(
        self,
        min_row: int,
        min_col: int | str,
        max_row: int,
        max_col: int | str,
        sheet: str | None = None,
    ) -> None:
        """Merge a range of cells."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        min_c = column_index_from_string(min_col) if isinstance(min_col, str) else min_col
        max_c = column_index_from_string(max_col) if isinstance(max_col, str) else max_col
        ws.merge_cells(
            start_row=min_row, start_column=min_c, end_row=max_row, end_column=max_c
        )

    def unmerge_cells(
        self,
        min_row: int,
        min_col: int | str,
        max_row: int,
        max_col: int | str,
        sheet: str | None = None,
    ) -> None:
        """Unmerge a previously merged cell range."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        min_c = column_index_from_string(min_col) if isinstance(min_col, str) else min_col
        max_c = column_index_from_string(max_col) if isinstance(max_col, str) else max_col
        ws.unmerge_cells(
            start_row=min_row, start_column=min_c, end_row=max_row, end_column=max_c
        )

    # ------------------------------------------------------------------
    # Row / column operations
    # ------------------------------------------------------------------

    def insert_rows(self, row: int, amount: int = 1, sheet: str | None = None) -> None:
        """Insert blank rows before the given row index."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        ws.insert_rows(row, amount)

    def delete_rows(self, row: int, amount: int = 1, sheet: str | None = None) -> None:
        """Delete rows starting at the given row index."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        ws.delete_rows(row, amount)

    def insert_cols(self, col: int | str, amount: int = 1, sheet: str | None = None) -> None:
        """Insert blank columns before the given column."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.insert_cols(col_idx, amount)

    def delete_cols(self, col: int | str, amount: int = 1, sheet: str | None = None) -> None:
        """Delete columns starting at the given column."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        col_idx = column_index_from_string(col) if isinstance(col, str) else col
        ws.delete_cols(col_idx, amount)

    # ------------------------------------------------------------------
    # Dimensions
    # ------------------------------------------------------------------

    def get_dimensions(self, sheet: str | None = None) -> tuple[int, int]:
        """Return (max_row, max_col) for the used area of a sheet."""
        ws = self.workbook[sheet] if sheet else self.workbook.active
        return ws.max_row, ws.max_column

    # ------------------------------------------------------------------
    # Save / load
    # ------------------------------------------------------------------

    def save(self, file_path: str | Path | None = None) -> None:
        """Save the workbook. Uses the path provided at init if none given."""
        path = Path(file_path) if file_path else self.file_path
        if path is None:
            raise ValueError("No file path specified for save().")
        self.workbook.save(path)
        self.file_path = path

    def load(self, file_path: str | Path) -> None:
        """Load a workbook from disk, replacing the current one."""
        self.file_path = Path(file_path)
        self.workbook = openpyxl.load_workbook(self.file_path)
