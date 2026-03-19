import datetime
import pytest
import openpyxl


def _create_valid_oc_workbook(path):
    """Create an Excel file with a 'Cadastro' sheet (v1) and an 'OC' sheet with sample data."""
    wb = openpyxl.Workbook()
    # Cadastro sheet with version v1
    cadastro = wb.active
    cadastro.title = "Cadastro"
    cadastro.cell(row=1, column=1, value="v1")

    # OC sheet with header and two data rows
    oc = wb.create_sheet("OC")
    header = ["Unidade", "ID Fornecedor", "Valor", "Vencimento", "Observação"]
    for col, h in enumerate(header, start=1):
        oc.cell(row=1, column=col, value=h)

    oc.cell(row=2, column=1, value="Belo Horizonte")
    oc.cell(row=2, column=2, value=123)
    oc.cell(row=2, column=3, value=1500.00)
    oc.cell(row=2, column=4, value=datetime.datetime(2026, 4, 30))
    oc.cell(row=2, column=5, value="Observação teste")

    oc.cell(row=3, column=1, value="Santo Amaro")
    oc.cell(row=3, column=2, value=456)
    oc.cell(row=3, column=3, value=800.00)
    oc.cell(row=3, column=4, value=datetime.datetime(2026, 5, 15))
    oc.cell(row=3, column=5, value="Outra observação")

    wb.save(path)
    return path


@pytest.fixture
def valid_oc_file(tmp_path):
    """Valid Excel file with a 'Cadastro' sheet (v1) and an 'OC' sheet."""
    return _create_valid_oc_workbook(tmp_path / "oc_valid.xlsx")


@pytest.fixture
def invalid_oc_file_missing_sheet(tmp_path):
    """Excel file without a 'Cadastro' sheet."""
    wb = openpyxl.Workbook()
    wb.active.title = "OC"
    path = tmp_path / "missing_sheet.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def invalid_oc_file_wrong_version(tmp_path):
    """Excel file with a 'Cadastro' sheet but version 'v2'."""
    wb = openpyxl.Workbook()
    cadastro = wb.active
    cadastro.title = "Cadastro"
    cadastro.cell(row=1, column=1, value="v2")
    wb.create_sheet("OC")
    path = tmp_path / "wrong_version.xlsx"
    wb.save(path)
    return path
