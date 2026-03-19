import datetime
import pytest

from utils.oc_loader import (
    FileValidationError,
    generate_purchase_order,
    generate_purchase_order_bulk,
    load_oc,
    validate_file,
)
from m8.purchase_order import PurchaseOrder


# ---------------------------------------------------------------------------
# validate_file
# ---------------------------------------------------------------------------

def test_validate_file_valid(valid_oc_file):
    # Should not raise
    validate_file(valid_oc_file)


def test_validate_file_missing_sheet(invalid_oc_file_missing_sheet):
    with pytest.raises(FileValidationError):
        validate_file(invalid_oc_file_missing_sheet)


def test_validate_file_wrong_version(invalid_oc_file_wrong_version):
    with pytest.raises(FileValidationError):
        validate_file(invalid_oc_file_wrong_version)


# ---------------------------------------------------------------------------
# load_oc
# ---------------------------------------------------------------------------

def test_load_oc_returns_expected_columns(valid_oc_file):
    expected_columns = {"Unidade", "ID Fornecedor", "Valor", "Vencimento", "Observação"}
    rows = load_oc(valid_oc_file)
    assert len(rows) > 0
    assert expected_columns.issubset(rows[0].keys())


def test_load_oc_skips_empty_rows(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OC"
    header = ["Unidade", "ID Fornecedor", "Valor", "Vencimento", "Observação"]
    for col, h in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=h)
    # Data row
    ws.cell(row=2, column=1, value="Contagem")
    ws.cell(row=2, column=2, value=99)
    ws.cell(row=2, column=3, value=500.0)
    ws.cell(row=2, column=4, value=datetime.datetime(2026, 6, 1))
    ws.cell(row=2, column=5, value="Obs")
    # Row 3 is completely empty
    path = tmp_path / "with_empty_row.xlsx"
    wb.save(path)

    rows = load_oc(path)
    assert len(rows) == 1


# ---------------------------------------------------------------------------
# generate_purchase_order
# ---------------------------------------------------------------------------

DUE_DATE = "2026-04-30T00:00:00Z"


def test_sp_unit_generates_company_id_1():
    po = generate_purchase_order("Santo Amaro", 123, 1000.0, DUE_DATE, "obs")
    assert po.empresaId == 1


def test_mg_unit_generates_company_id_31():
    po = generate_purchase_order("Belo Horizonte", 123, 1000.0, DUE_DATE, "obs")
    assert po.empresaId == 31


def test_employee_id_derived_from_company():
    po_sp = generate_purchase_order("Guarulhos", 1, 100.0, DUE_DATE, "obs")
    assert po_sp.funcionarioId == 1411093179

    po_mg = generate_purchase_order("Uberlândia", 1, 100.0, DUE_DATE, "obs")
    assert po_mg.funcionarioId == 1411091897


def test_cost_center_derived_from_unit():
    po_sp = generate_purchase_order("Carapicuíba", 1, 100.0, DUE_DATE, "obs")
    assert po_sp.items[0].centroCustoId == 0

    po_bh = generate_purchase_order("Belo Horizonte", 1, 100.0, DUE_DATE, "obs")
    assert po_bh.items[0].centroCustoId == 287

    po_contagem = generate_purchase_order("Contagem", 1, 100.0, DUE_DATE, "obs")
    assert po_contagem.items[0].centroCustoId == 286


def test_fiscal_operation_derived_from_company():
    po_sp = generate_purchase_order("Sacomã", 1, 100.0, DUE_DATE, "obs")
    assert po_sp.items[0].operacaoFiscalId == 1

    po_mg = generate_purchase_order("Montes Claros", 1, 100.0, DUE_DATE, "obs")
    assert po_mg.items[0].operacaoFiscalId == 250


def test_hardcoded_defaults():
    po = generate_purchase_order("Diadema", 42, 999.0, DUE_DATE, "obs")
    assert po.status == "Aprovado"
    assert po.freteId == 9
    assert po.condicaoPagamentoId == 15
    assert po.tipoOrdemCompraId == 1
    assert po.items[0].produtoId == 2307
    assert po.items[0].unidadeId == 5
    assert po.items[0].quantidade == 1


def test_returns_purchase_order_with_one_item_and_one_installment():
    po = generate_purchase_order("Venda Nova", 7, 250.0, DUE_DATE, "obs")
    assert isinstance(po, PurchaseOrder)
    assert len(po.items) == 1
    assert len(po.installments) == 1


# ---------------------------------------------------------------------------
# generate_purchase_order_bulk
# ---------------------------------------------------------------------------

def test_generate_purchase_order_bulk_returns_same_length_list():
    items = [
        {
            "Unidade": "Belo Horizonte",
            "ID Fornecedor": 10,
            "Valor": 500.0,
            "Vencimento": datetime.datetime(2026, 4, 1),
            "Observação": "obs 1",
        },
        {
            "Unidade": "Santo Amaro",
            "ID Fornecedor": 20,
            "Valor": 750.0,
            "Vencimento": datetime.datetime(2026, 5, 1),
            "Observação": "obs 2",
        },
        {
            "Unidade": "Contagem",
            "ID Fornecedor": 30,
            "Valor": 1200.0,
            "Vencimento": datetime.datetime(2026, 6, 1),
            "Observação": "obs 3",
        },
    ]
    result = generate_purchase_order_bulk(items)
    assert len(result) == len(items)
    assert all(isinstance(po, PurchaseOrder) for po in result)
